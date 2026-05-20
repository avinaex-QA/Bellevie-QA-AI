import io
import json

from openpyxl import load_workbook

from backend.models.schemas import TestCase
from backend.services import excel_service


def test_excel_export_includes_project_context_and_history(tmp_path, monkeypatch):
    monkeypatch.setattr(excel_service, "PROJECT_ROOT", tmp_path)
    monkeypatch.setattr(excel_service, "EXPORTS_DIR", tmp_path / "exports")
    monkeypatch.setattr(excel_service, "HISTORY_FILE", tmp_path / "exports" / "history.json")

    raw_bytes, file_name = excel_service.generate_excel(
        test_cases=[
            TestCase(
                id="TC-001",
                priority="High",
                title="Verify login",
                preconditions="User exists",
                steps=["Open app", "Log in"],
                expected_result="User logs in",
                tags=["Smoke"],
                test_type="Functional",
            )
        ],
        selected_projects=["Resident APP", "VMS APP"],
        selected_modules=["Onboarding", "VMS"],
        source_type="text",
        module_detected="Login",
    )

    workbook = load_workbook(io.BytesIO(raw_bytes))
    info_sheet = workbook["Project Info"]
    metadata = {info_sheet.cell(row=row, column=1).value: info_sheet.cell(row=row, column=2).value for row in range(2, 12)}

    assert metadata["Project Context"] == "Resident APP, VMS APP"
    assert metadata["Module Context"] == "Onboarding, VMS"
    assert (tmp_path / "exports" / file_name).exists()

    history = json.loads((tmp_path / "exports" / "history.json").read_text(encoding="utf-8"))
    assert history["exports"][0]["selected_projects"] == ["Resident APP", "VMS APP"]
    assert history["exports"][0]["selected_modules"] == ["Onboarding", "VMS"]

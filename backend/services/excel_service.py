"""
Excel export service.
Generates a formatted, colour-coded .xlsx file, saves it to /exports/,
updates history.json, and returns raw bytes for streaming download.
"""
import io
import json
import os
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.config.settings import settings
from backend.models.schemas import TestCase
from backend.utils.logger import setup_logger

logger = setup_logger(__name__)

# ── Paths ──────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
EXPORTS_DIR  = PROJECT_ROOT / "exports"
HISTORY_FILE = EXPORTS_DIR / "history.json"

# ── Colour palette ─────────────────────────────────────────────────────────
HEADER_BG = "1E3A5F"
HEADER_FG = "FFFFFF"
HIGH_BG   = "FF4D4D"
HIGH_FG   = "FFFFFF"
MED_BG    = "FFB347"
MED_FG    = "1A1A1A"
LOW_BG    = "4CAF50"
LOW_FG    = "FFFFFF"
ROW_ALT   = "F0F4FA"
ROW_NORM  = "FFFFFF"

# ── Column layout ──────────────────────────────────────────────────────────
COLUMNS = [
    ("Test Case ID",       12),
    ("Priority",           10),
    ("Test Type",          14),
    ("Tags",               18),
    ("Title",              45),
    ("Preconditions",      35),
    ("Steps to Reproduce", 55),
    ("Expected Result",    45),
    ("Actual Result",      35),
]
HEADERS    = [c[0] for c in COLUMNS]
COL_WIDTHS = [c[1] for c in COLUMNS]


# ── Style helpers ──────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    thin = Side(style="thin", color="C0C8D8")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _priority_colors(priority: str) -> tuple[str, str]:
    p = priority.strip().lower()
    if p == "high":   return HIGH_BG, HIGH_FG
    if p == "medium": return MED_BG, MED_FG
    return LOW_BG, LOW_FG


def _format_steps(steps: List[str]) -> str:
    return "\n".join(f"{i}. {step}" for i, step in enumerate(steps, 1))


# ── History helpers ────────────────────────────────────────────────────────
def load_history() -> dict:
    if HISTORY_FILE.exists():
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"exports": []}


def save_history(data: dict) -> None:
    try:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        logger.warning(f"Could not write history.json: {e}")


def append_history_entry(
    file_name: str,
    file_path: Path,
    count: int,
    source: str = "text",
    module: str = "General",
    selected_projects: Optional[List[str]] = None,
) -> None:
    """Prepend a new export entry to history.json (newest first)."""
    history = load_history()
    history["exports"].insert(0, {
        "file_name": file_name,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "count": count,
        "path": str(file_path.relative_to(PROJECT_ROOT)).replace("\\", "/"),
        "source": source,
        "module": module,
        "selected_projects": selected_projects or [],
    })
    save_history(history)
    logger.info(f"History updated — {len(history['exports'])} total exports recorded")


def remove_history_entry(file_name: str) -> None:
    """Remove a single entry from history.json by file name."""
    history = load_history()
    history["exports"] = [e for e in history["exports"] if e["file_name"] != file_name]
    save_history(history)


# ── Auto-open helper ───────────────────────────────────────────────────────
def _auto_open_file(file_path: Path) -> None:
    """
    Open the saved Excel file using the OS default application.

    Rules:
    - Only runs when AUTO_OPEN_EXCEL=true in .env
    - Only runs on Windows (os.startfile is a Windows-only API)
    - Runs ONCE per export, immediately after the file is fully written to disk
    - Non-blocking: os.startfile returns instantly; Excel opens in the background
    - Any error is logged as a warning and silently ignored so the API response
      is never affected
    """
    if not settings.auto_open_excel:
        return  # feature is disabled — skip silently

    if os.name != "nt":
        logger.info("Auto-open skipped: os.startfile is only supported on Windows.")
        return

    try:
        os.startfile(str(file_path))
        logger.info(f"Auto-open triggered: {file_path.name}")
    except OSError as e:
        logger.warning(f"Auto-open failed (file={file_path.name}): {e}")


# ── Main Excel generator ───────────────────────────────────────────────────
def generate_excel(
    test_cases: List[TestCase],
    selected_projects: Optional[List[str]] = None,
    project_name: str = "AI Generated Test Cases",
    sheet_title: str = "Test Cases",
    source_type: str = "text",
    module_detected: str = "General",
) -> tuple[bytes, str]:
    """
    Builds a formatted workbook, saves it to /exports/<timestamp>.xlsx,
    updates history.json, and returns (raw_bytes, file_name).
    """
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    selected_projects = selected_projects or []
    project_context = ", ".join(selected_projects) if selected_projects else "Not specified"

    # ── Timestamp filename ────────────────────────────────────────────────
    ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"test_cases_{ts}.xlsx"
    file_path = EXPORTS_DIR / file_name

    wb = Workbook()

    # ── Project Info sheet ────────────────────────────────────────────────
    info_ws       = wb.active
    info_ws.title = "Project Info"
    info_ws.sheet_view.showGridLines = False
    info_ws.column_dimensions["A"].width = 25
    info_ws.column_dimensions["B"].width = 45

    high   = sum(1 for t in test_cases if t.priority.lower() == "high")
    medium = sum(1 for t in test_cases if t.priority.lower() == "medium")
    low    = sum(1 for t in test_cases if t.priority.lower() == "low")

    info_data = [
        ("Project Name",    project_name),
        ("Generated By",    "AI Test Case Generator"),
        ("Generated On",    datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Project Context", project_context),
        ("Source",          source_type),
        ("Module",          module_detected),
        ("Total Test Cases", len(test_cases)),
        ("High Priority",   high),
        ("Medium Priority", medium),
        ("Low Priority",    low),
    ]
    for row_idx, (label, value) in enumerate(info_data, start=2):
        lc = info_ws.cell(row=row_idx, column=1, value=label)
        vc = info_ws.cell(row=row_idx, column=2, value=str(value))
        lc.font   = Font(bold=True, color=HEADER_BG)
        lc.fill   = _fill("EEF2FF")
        lc.border = _border()
        vc.border = _border()
        vc.alignment = Alignment(wrap_text=False)

    # ── Test Cases sheet ──────────────────────────────────────────────────
    tc_ws              = wb.create_sheet(title=sheet_title)
    tc_ws.sheet_view.showGridLines = False
    tc_ws.freeze_panes = "A2"

    for col_idx, header in enumerate(HEADERS, start=1):
        cell            = tc_ws.cell(row=1, column=col_idx, value=header)
        cell.font       = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
        cell.fill       = _fill(HEADER_BG)
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = _border()
        tc_ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]
    tc_ws.row_dimensions[1].height = 28

    for row_idx, tc in enumerate(test_cases, start=2):
        bg     = _fill(ROW_ALT if row_idx % 2 == 0 else ROW_NORM)
        p_bg, p_fg = _priority_colors(tc.priority)

        row_data = [
            tc.id,
            tc.priority,
            tc.test_type,
            ", ".join(tc.tags),
            tc.title,
            tc.preconditions,
            _format_steps(tc.steps),
            tc.expected_result,
            tc.actual_result,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell           = tc_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Calibri", size=10)
            cell.border    = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            if col_idx == 2:    # Priority — coloured
                cell.fill      = _fill(p_bg)
                cell.font      = Font(name="Calibri", bold=True, color=p_fg, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 1:  # ID
                cell.fill      = _fill("E8EDFF")
                cell.font      = Font(name="Calibri", bold=True, color=HEADER_BG, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.fill = bg

        step_count = len(tc.steps)
        tc_ws.row_dimensions[row_idx].height = min(max(40, 15 * (step_count + 1)), 200)

    tc_ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    # ── Save to disk ──────────────────────────────────────────────────────
    wb.save(str(file_path))
    logger.info(f"Excel saved at: {file_path}")

    # ── Auto-open (optional, Windows only, controlled by AUTO_OPEN_EXCEL) ─
    # Called exactly once per export, after the file is fully written.
    # os.startfile is non-blocking — it returns immediately so the API
    # response is never delayed.
    _auto_open_file(file_path)

    # ── Also capture as bytes for HTTP streaming ──────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    raw_bytes = output.getvalue()

    # ── Update history ────────────────────────────────────────────────────
    append_history_entry(
        file_name=file_name,
        file_path=file_path,
        count=len(test_cases),
        source=source_type,
        module=module_detected,
        selected_projects=selected_projects,
    )

    return raw_bytes, file_name
